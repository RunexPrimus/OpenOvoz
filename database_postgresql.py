import psycopg2
import psycopg2.extras
import json
from datetime import datetime
from config import (
    POSTGRES_HOST, POSTGRES_PORT, POSTGRES_DB, 
    POSTGRES_USER, POSTGRES_PASSWORD, DATABASE_URL, DATABASE_PUBLIC_URL
)

class DatabasePostgreSQL:
    def __init__(self):
        self.connection_params = self._get_connection_params()
        self.init_database()
    
    def _get_connection_params(self):
        """PostgreSQL ulanish parametrlarini olish"""
        # Railway da avval DATABASE_PUBLIC_URL ni sinab ko'rish (eng muhim - tashqi ulanish uchun)
        if DATABASE_PUBLIC_URL:
            print(f"Railway DATABASE_PUBLIC_URL ishlatilmoqda: {DATABASE_PUBLIC_URL[:50]}...")
            return {'dsn': DATABASE_PUBLIC_URL}
        elif DATABASE_URL:
            print(f"Railway DATABASE_URL ishlatilmoqda: {DATABASE_URL[:50]}...")
            # DATABASE_URL ni parse qilib, individual parametrlarga o'tkazish
            try:
                from urllib.parse import urlparse
                parsed = urlparse(DATABASE_URL)
                # SSL rejimini aniqroq sozlash
                ssl_mode = 'require' if 'railway' in DATABASE_URL.lower() else 'prefer'
                
                return {
                    'host': parsed.hostname,
                    'port': parsed.port or 5432,
                    'database': parsed.path[1:],  # /database_name -> database_name
                    'user': parsed.username,
                    'password': parsed.password,
                    'sslmode': ssl_mode,
                    'connect_timeout': 30,
                    'application_name': 'BotopneBot'
                }
            except Exception as e:
                print(f"DATABASE_URL parse qilishda xato: {e}")
                # Fallback uchun individual parametrlar
                return {
                    'host': POSTGRES_HOST,
                    'port': POSTGRES_PORT,
                    'database': POSTGRES_DB,
                    'user': POSTGRES_USER,
                    'password': POSTGRES_PASSWORD,
                    'sslmode': 'prefer',
                    'connect_timeout': 30,
                    'application_name': 'BotopneBot'
                }
        else:
            # Fallback uchun individual parametrlar (faqat local development uchun)
            print(f"Fallback parametrlar ishlatilmoqda: {POSTGRES_HOST}:{POSTGRES_PORT}")
            return {
                'host': POSTGRES_HOST,
                'port': POSTGRES_PORT,
                'database': POSTGRES_DB,
                'user': POSTGRES_USER,
                'password': POSTGRES_PASSWORD,
                'sslmode': 'prefer',
                'connect_timeout': 30,
                'application_name': 'BotopneBot'
            }
    
    def get_connection(self):
        """PostgreSQL ulanishini olish"""
        max_retries = 5
        retry_delay = 1  # sekund
        
        for attempt in range(max_retries):
            try:
                print(f"Ulanish urinishi {attempt + 1}/{max_retries}...")
                print(f"Connection params: host={self.connection_params.get('host', 'N/A')}, port={self.connection_params.get('port', 'N/A')}, database={self.connection_params.get('database', 'N/A')}")
                
                conn = psycopg2.connect(**self.connection_params)
                conn.autocommit = False
                print("PostgreSQL ulanish muvaffaqiyatli!")
                
                # Connection ni test qilish
                cursor = conn.cursor()
                cursor.execute("SELECT 1")
                cursor.fetchone()
                cursor.close()
                print("PostgreSQL connection test muvaffaqiyatli!")
                
                return conn
            except psycopg2.OperationalError as e:
                print(f"PostgreSQL ulanishda xato (urinish {attempt + 1}): {e}")
                if attempt < max_retries - 1:
                    print(f"{retry_delay} soniyadan keyin qayta urinish...")
                    import time
                    time.sleep(retry_delay)
                    retry_delay = min(retry_delay * 2, 30)  # Exponential backoff, max 30 seconds
                else:
                    print("Barcha urinishlar muvaffaqiyatsiz tugadi!")
                    print(f"Final error: {e}")
                    raise
            except psycopg2.Error as e:
                print(f"PostgreSQL xatosi: {e}")
                print(f"Xato kodi: {e.pgcode if hasattr(e, 'pgcode') else 'N/A'}")
                raise
            except Exception as e:
                print(f"Kutilmagan xato: {e}")
                print(f"Xato turi: {type(e).__name__}")
                raise
    
    def init_database(self):
        """Ma'lumotlar bazasi va jadvallarni yaratish"""
        conn = None
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            print("Ma'lumotlar bazasi jadvallari yaratilmoqda...")
            
            # Foydalanuvchilar jadvali
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    id SERIAL PRIMARY KEY,
                    telegram_id BIGINT UNIQUE NOT NULL,
                    username VARCHAR(255),
                    first_name VARCHAR(255),
                    last_name VARCHAR(255),
                    phone VARCHAR(20),
                    region VARCHAR(100),
                    language VARCHAR(10) DEFAULT 'uz',
                    referral_code VARCHAR(20) UNIQUE,
                    referred_by INTEGER,
                    balance INTEGER DEFAULT 0,
                    pending_balance INTEGER DEFAULT 0,
                    total_earned INTEGER DEFAULT 0,
                    total_withdrawn INTEGER DEFAULT 0,
                    last_withdrawal_account TEXT,
                    is_active BOOLEAN DEFAULT TRUE,
                    is_admin BOOLEAN DEFAULT FALSE,
                    is_super_admin BOOLEAN DEFAULT FALSE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            print("✅ users jadvali yaratildi/yangilandi")
            
            # Mavsumlar jadvali
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS seasons (
                    id SERIAL PRIMARY KEY,
                    name VARCHAR(255) NOT NULL,
                    description TEXT,
                    start_date DATE NOT NULL,
                    end_date DATE NOT NULL,
                    max_votes_per_user INTEGER DEFAULT 3,
                    is_active BOOLEAN DEFAULT FALSE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            print("✅ seasons jadvali yaratildi/yangilandi")
            
            # Loyihalar jadvali
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS projects (
                    id SERIAL PRIMARY KEY,
                    season_id INTEGER NOT NULL,
                    name VARCHAR(255) NOT NULL,
                    description TEXT,
                    budget INTEGER NOT NULL,
                    region VARCHAR(100) NOT NULL,
                    category VARCHAR(100),
                    image_url TEXT,
                    project_url TEXT,
                    link TEXT,
                    status VARCHAR(50) DEFAULT 'draft',
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (season_id) REFERENCES seasons (id)
                )
            ''')
            print("✅ projects jadvali yaratildi/yangilandi")
            
            # Ovozlar jadvali
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS votes (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL,
                    project_id INTEGER NOT NULL,
                    season_id INTEGER NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users (id),
                    FOREIGN KEY (project_id) REFERENCES projects (id),
                    FOREIGN KEY (season_id) REFERENCES seasons (id),
                    UNIQUE(user_id, project_id, season_id)
                )
            ''')
            print("✅ votes jadvali yaratildi/yangilandi")
            
            # Pul chiqarish so'rovlari
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS withdrawal_requests (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL,
                    amount INTEGER NOT NULL,
                    commission INTEGER NOT NULL,
                    net_amount INTEGER NOT NULL,
                    method VARCHAR(50) NOT NULL,
                    account_details TEXT NOT NULL,
                    status VARCHAR(50) DEFAULT 'pending',
                    admin_notes TEXT,
                    processed_by INTEGER,
                    processed_at TIMESTAMP,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users (id),
                    FOREIGN KEY (processed_by) REFERENCES users (id)
                )
            ''')
            print("✅ withdrawal_requests jadvali yaratildi/yangilandi")
            
            # Sozlamalar jadvali
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS settings (
                    id SERIAL PRIMARY KEY,
                    key VARCHAR(100) UNIQUE NOT NULL,
                    value TEXT NOT NULL,
                    description TEXT,
                    updated_by INTEGER,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (updated_by) REFERENCES users (id)
                )
            ''')
            print("✅ settings jadvali yaratildi/yangilandi")
            
            # Balans tarixi
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS balance_history (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL,
                    amount INTEGER NOT NULL,
                    type VARCHAR(100) NOT NULL,
                    description TEXT,
                    reference_id INTEGER,
                    status VARCHAR(50) DEFAULT 'pending',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users (id)
                )
            ''')
            print("✅ balance_history jadvali yaratildi/yangilandi")
            
            # Yangiliklar
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS announcements (
                    id SERIAL PRIMARY KEY,
                    title VARCHAR(255) NOT NULL,
                    content TEXT NOT NULL,
                    language VARCHAR(10) DEFAULT 'uz',
                    is_active BOOLEAN DEFAULT TRUE,
                    created_by INTEGER NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (created_by) REFERENCES users (id)
                )
            ''')
            print("✅ announcements jadvali yaratildi/yangilandi")
            
            # Audit loglari
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS audit_logs (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER,
                    action VARCHAR(255) NOT NULL,
                    details TEXT,
                    ip_address VARCHAR(45),
                    user_agent TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users (id)
                )
            ''')
            print("✅ audit_logs jadvali yaratildi/yangilandi")
            
            # Tasdiqlangan loyihalar jadvali
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS approved_projects (
                    id SERIAL PRIMARY KEY,
                    name VARCHAR(255) NOT NULL,
                    link TEXT NOT NULL,
                    status VARCHAR(50) DEFAULT 'approved',
                    approved_by INTEGER NOT NULL,
                    approved_at TIMESTAMP NOT NULL,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (approved_by) REFERENCES users (id)
                )
            ''')
            print("✅ approved_projects jadvali yaratildi/yangilandi")
            
            # Dastlabki sozlamalarni qo'shish
            cursor.execute('''
                INSERT INTO settings (key, value, description) VALUES 
                ('REFERRAL_BONUS', '1000', 'Referal uchun bonus puli'),
                ('VOTE_BONUS', '25000', 'Ovoz berish uchun bonus puli'),
                ('MIN_WITHDRAWAL', '20000', 'Minimal yechish miqdori'),
                ('COMMISSION_RATE', '0.02', 'Komissiya foizi (0.01 = 1%)')
                ON CONFLICT (key) DO NOTHING
            ''')
            print("✅ Dastlabki sozlamalar qo'shildi")
            
            conn.commit()
            print("PostgreSQL jadvallari muvaffaqiyatli yaratildi!")
            
        except Exception as e:
            if conn:
                conn.rollback()
            print(f"PostgreSQL jadvallarini yaratishda xato: {e}")
            raise
        finally:
            if conn:
                conn.close()
    
    def create_user(self, telegram_id, username, first_name, last_name, phone, region, language, referred_by=None):
        """Yangi foydalanuvchi yaratish"""
        conn = self.get_connection()
        try:
            # Referal ID yaratish (8 ta belgi)
            import secrets
            import string
            
            def generate_referral_id():
                """Benzersiz referal ID yaratish"""
                while True:
                    # 8 ta belgi: harflar va raqamlar
                    chars = string.ascii_uppercase + string.digits
                    referral_id = ''.join(secrets.choice(chars) for _ in range(8))
                    
                    # Bu ID mavjud emasligini tekshirish
                    cursor = conn.cursor()
                    cursor.execute("SELECT id FROM users WHERE referral_code = %s", (referral_id,))
                    if not cursor.fetchone():
                        return referral_id
            
            referral_code = generate_referral_id()
            
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO users (telegram_id, username, first_name, last_name, phone, region, language, referred_by, referral_code, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (telegram_id, username, first_name, last_name, phone, region, language, referred_by, referral_code, datetime.now()))
            
            user_id = cursor.fetchone()[0]
            
            # Agar referal orqali kelgan bo'lsa, bonus berish
            if referred_by:
                self.add_balance(referred_by, 1000, 'referral_bonus', f'Yangi foydalanuvchi jalb etildi')
                self.add_balance(user_id, 1000, 'referral_bonus', f'Referal orqali ro\'yxatdan o\'tish')
            
            conn.commit()
            return user_id
        except Exception as e:
            conn.rollback()
            print(f"Foydalanuvchi yaratishda xato: {e}")
            return None
        finally:
            conn.close()
    
    def get_user(self, telegram_id):
        """Foydalanuvchini telegram_id bo'yicha olish"""
        conn = self.get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        try:
            cursor.execute('SELECT * FROM users WHERE telegram_id = %s', (telegram_id,))
            user = cursor.fetchone()
            return dict(user) if user else None
        finally:
            cursor.close()
            conn.close()
    
    def add_balance(self, user_id, amount, balance_type, description):
        """Foydalanuvchi balansiga qo'shish"""
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            
            # Balansni yangilash
            cursor.execute("""
                UPDATE users SET 
                balance = balance + %s,
                total_earned = total_earned + %s,
                updated_at = %s
                WHERE id = %s
            """, (amount, amount, datetime.now(), user_id))
            
            # Balans tarixiga qo'shish
            cursor.execute("""
                INSERT INTO balance_history (user_id, amount, type, description, status, created_at)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (user_id, amount, balance_type, description, 'approved', datetime.now()))
            
            conn.commit()
            return True
        except Exception as e:
            conn.rollback()
            print(f"Balans qo'shishda xato: {e}")
            return False
        finally:
            conn.close()
    
    def get_active_season(self):
        """Faol mavsumni olish"""
        conn = self.get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        try:
            cursor.execute('''
                SELECT * FROM seasons 
                WHERE is_active = TRUE AND CURRENT_DATE BETWEEN start_date AND end_date
                ORDER BY start_date DESC LIMIT 1
            ''')
            
            season = cursor.fetchone()
            return dict(season) if season else None
        finally:
            cursor.close()
            conn.close()
    
    def get_projects_by_season(self, season_id, region=None):
        """Mavsum bo'yicha loyihalarni olish"""
        conn = self.get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        try:
            if region:
                cursor.execute('''
                    SELECT * FROM projects 
                    WHERE season_id = %s AND region = %s AND status = 'published'
                    ORDER BY name
                ''', (season_id, region))
            else:
                cursor.execute('''
                    SELECT * FROM projects 
                    WHERE season_id = %s AND status = 'published'
                    ORDER BY name
                ''', (season_id,))
            
            projects = cursor.fetchall()
            return [dict(project) for project in projects]
        finally:
            cursor.close()
            conn.close()
    
    def get_approved_projects(self):
        """Tasdiqlangan loyihalarni olish"""
        conn = self.get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        try:
            cursor.execute("""
                SELECT id, name, link, status, approved_by, approved_at, created_at
                FROM approved_projects 
                WHERE status = 'approved'
                ORDER BY created_at DESC
            """)
            
            projects = cursor.fetchall()
            return [dict(project) for project in projects]
        finally:
            cursor.close()
            conn.close()
    
    def get_setting(self, key, default=None):
        """Sozlama qiymatini olish"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute("SELECT value FROM settings WHERE key = %s", (key,))
            result = cursor.fetchone()
            return result[0] if result else default
        finally:
            cursor.close()
            conn.close()
    
    def update_setting(self, key, value, admin_id):
        """Sozlama qiymatini yangilash"""
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            
            # Avval sozlama mavjudligini tekshirish
            cursor.execute("SELECT id FROM settings WHERE key = %s", (key,))
            if not cursor.fetchone():
                print(f"Sozlama topilmadi: {key}")
                return False
            
            # Sozlamani yangilash
            cursor.execute("""
                UPDATE settings 
                SET value = %s, updated_at = CURRENT_TIMESTAMP 
                WHERE key = %s
            """, (value, key))
            
            if cursor.rowcount > 0:
                conn.commit()
                print(f"Sozlama muvaffaqiyatli yangilandi: {key} = {value}")
                return True
            else:
                print(f"Sozlama yangilanmadi: {key}")
                return False
                
        except Exception as e:
            conn.rollback()
            print(f"Sozlama yangilashda xato: {e}")
            return False
        finally:
            conn.close()
    
    def get_statistics(self):
        """Umumiy statistikalarni olish"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            # Foydalanuvchilar soni
            cursor.execute('SELECT COUNT(*) as total_users FROM users WHERE is_active = TRUE')
            total_users = cursor.fetchone()[0]
            
            # Bugungi yangi foydalanuvchilar
            cursor.execute('''
                SELECT COUNT(*) as today_users 
                FROM users 
                WHERE DATE(created_at) = CURRENT_DATE AND is_active = TRUE
            ''')
            today_users = cursor.fetchone()[0]
            
            # Jami ovozlar - to'lov qilingan foydalanuvchilar soni
            cursor.execute('SELECT COUNT(DISTINCT user_id) as total_votes FROM balance_history WHERE type = %s AND status = %s', ('payment', 'approved'))
            total_votes = cursor.fetchone()[0]
            
            # Jami balans - tasdiqlangan pullar summasi
            cursor.execute('SELECT COALESCE(SUM(amount), 0) as total_balance FROM balance_history WHERE type = %s AND status = %s', ('payment', 'approved'))
            total_balance = cursor.fetchone()[0] or 0
            
            return {
                'total_users': total_users,
                'today_users': today_users,
                'total_votes': total_votes,
                'total_balance': total_balance
            }
        finally:
            cursor.close()
            conn.close()
    
    def get_referral_stats(self, user_id):
        """Foydalanuvchining referal statistikalarini olish"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            # Jalb etilgan foydalanuvchilar soni
            cursor.execute("SELECT COUNT(*) FROM users WHERE referred_by = %s", (user_id,))
            referrals_count = cursor.fetchone()[0]
            
            # Referal orqali olingan bonuslar
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0) FROM balance_history 
                WHERE user_id = %s AND type = 'referral_bonus'
            """, (user_id,))
            referral_earnings = cursor.fetchone()[0] or 0
            
            return {
                'referrals_count': referrals_count,
                'referral_earnings': referral_earnings
            }
        finally:
            cursor.close()
            conn.close()
    
    def get_user_by_referral_code(self, referral_code):
        """Referal kod orqali foydalanuvchini topish"""
        conn = self.get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        try:
            cursor.execute("""
                SELECT id, telegram_id, username, first_name, last_name, phone, region, language, 
                       referral_code, referred_by, balance, pending_balance, total_withdrawn, 
                       total_earned, created_at, updated_at
                FROM users WHERE referral_code = %s
            """, (referral_code,))
            
            row = cursor.fetchone()
            return dict(row) if row else None
        finally:
            cursor.close()
            conn.close()
    
    def get_all_active_users(self):
        """Barcha faol foydalanuvchilarni olish"""
        conn = self.get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        try:
            cursor.execute("""
                SELECT telegram_id, first_name, language
                FROM users
                WHERE is_active = TRUE
                ORDER BY created_at DESC
            """)
            users = cursor.fetchall()
            return [dict(user) for user in users]
        finally:
            cursor.close()
            conn.close()
    
    def create_approved_project(self, project_data):
        """Tasdiqlangan loyihani yaratish"""
        conn = None
        try:
            print(f"Loyiha yaratish boshlandi: {project_data}")
            
            # Database ulanishini tekshirish
            try:
                conn = self.get_connection()
                print("Database ulanish muvaffaqiyatli!")
            except Exception as e:
                print(f"Database ulanishda xato: {e}")
                return None
            
            cursor = conn.cursor()
            
            # Ma'lumotlarni tekshirish
            required_fields = ['name', 'link', 'status', 'approved_by', 'approved_at']
            for field in required_fields:
                if field not in project_data or project_data[field] is None:
                    print(f"Majburiy maydon topilmadi: {field}")
                    return None
            
            print(f"Loyiha ma'lumotlari to'g'ri, bazaga saqlash boshlandi...")
            
            # Jadval mavjudligini tekshirish
            cursor.execute("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_name = 'approved_projects'
                )
            """)
            
            if not cursor.fetchone()[0]:
                print("approved_projects jadvali mavjud emas!")
                return None
            
            print("approved_projects jadvali mavjud, INSERT boshlandi...")
            
            cursor.execute("""
                INSERT INTO approved_projects (name, link, status, approved_by, approved_at, created_at)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                project_data['name'],
                project_data['link'],
                project_data['status'],
                project_data['approved_by'],
                project_data['approved_at'],
                datetime.now()
            ))
            
            project_id = cursor.fetchone()[0]
            conn.commit()
            print(f"Loyiha muvaffaqiyatli yaratildi! ID: {project_id}")
            return project_id
            
        except Exception as e:
            if conn:
                conn.rollback()
            print(f"Tasdiqlangan loyiha yaratishda xato: {e}")
            print(f"Xato turi: {type(e).__name__}")
            import traceback
            print(f"Xato izi: {traceback.format_exc()}")
            return None
        finally:
            if conn:
                conn.close()
    
    def create_withdrawal_request(self, withdrawal_data):
        """Pul chiqarish so'rovini yaratish"""
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO withdrawal_requests (user_id, amount, commission, net_amount, method, account_details, status, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                withdrawal_data['user_id'],
                withdrawal_data['amount'],
                withdrawal_data['commission'],
                withdrawal_data['net_amount'],
                withdrawal_data['method'],
                withdrawal_data['account_details'],
                withdrawal_data['status'],
                datetime.now()
            ))
            
            withdrawal_id = cursor.fetchone()[0]
            
            # Foydalanuvchining balansini kamaytirish
            cursor.execute("""
                UPDATE users SET 
                balance = balance - %s,
                pending_balance = pending_balance + %s,
                last_withdrawal_account = %s,
                updated_at = %s
                WHERE id = %s
            """, (withdrawal_data['amount'], withdrawal_data['amount'], withdrawal_data['account_details'], datetime.now(), withdrawal_data['user_id']))
            
            conn.commit()
            return withdrawal_id
        except Exception as e:
            conn.rollback()
            print(f"Pul chiqarish so'rovini yaratishda xato: {e}")
            return None
        finally:
            conn.close()
    
    def complete_withdrawal(self, user_id, amount):
        """Pul chiqarishni yakunlash va balansni 0 qilish"""
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            
            # Balansni 0 qilish va total_withdrawn ni yangilash
            cursor.execute("""
                UPDATE users SET 
                balance = 0,
                pending_balance = 0,
                total_withdrawn = total_withdrawn + %s,
                updated_at = %s
                WHERE id = %s
            """, (amount, datetime.now(), user_id))
            
            # Withdrawal status ni 'completed' qilish
            cursor.execute("""
                UPDATE withdrawal_requests SET 
                status = 'completed',
                processed_at = %s
                WHERE user_id = %s AND status = 'pending'
            """, (datetime.now(), user_id))
            
            conn.commit()
            return True
        except Exception as e:
            conn.rollback()
            print(f"Pul chiqarishni yakunlashda xato: {e}")
            return False
        finally:
            conn.close()
    
    def reject_withdrawal(self, user_id):
        """Pul chiqarishni rad etish va balansni qaytarish"""
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            
            # Eng so'nggi pending withdrawal ni olish
            cursor.execute("""
                SELECT amount FROM withdrawal_requests 
                WHERE user_id = %s AND status = 'pending'
                ORDER BY created_at DESC
                LIMIT 1
            """, (user_id,))
            
            result = cursor.fetchone()
            if result:
                amount = result[0]
                
                # Balansni qaytarish
                cursor.execute("""
                    UPDATE users SET 
                    balance = balance + %s,
                    pending_balance = 0,
                    updated_at = %s
                    WHERE id = %s
                """, (amount, datetime.now(), user_id))
                
                # Withdrawal status ni 'rejected' qilish
                cursor.execute("""
                    UPDATE withdrawal_requests SET 
                    status = 'rejected',
                    processed_at = %s
                    WHERE user_id = %s AND status = 'pending'
                """, (datetime.now(), user_id))
                
                conn.commit()
                return True
            return False
        except Exception as e:
            conn.rollback()
            print(f"Pul chiqarishni rad etishda xato: {e}")
            return False
        finally:
            conn.close()
    
    def get_withdrawal_notification(self, withdrawal_id):
        """Pul chiqarish so'rovini ID bo'yicha olish"""
        conn = self.get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        try:
            cursor.execute("""
                SELECT 
                    wr.user_id,
                    wr.amount,
                    wr.method,
                    wr.account_details,
                    wr.processed_at,
                    wr.created_at,
                    u.telegram_id,
                    u.language
                FROM withdrawal_requests wr
                JOIN users u ON u.id = wr.user_id
                WHERE wr.id = %s
            """, (withdrawal_id,))
            
            result = cursor.fetchone()
            return dict(result) if result else None
        finally:
            cursor.close()
            conn.close()
    
    def get_top_voters(self, limit=10):
        """Eng ko'p ovoz bergan foydalanuvchilarni olish"""
        conn = self.get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        try:
            cursor.execute("""
                SELECT 
                    u.first_name,
                    u.telegram_id,
                    COUNT(bh.id) as vote_count
                FROM users u
                LEFT JOIN balance_history bh ON u.id = bh.user_id AND bh.type = 'payment' AND bh.status = 'approved'
                WHERE u.is_active = TRUE
                GROUP BY u.id, u.first_name, u.telegram_id
                HAVING COUNT(bh.id) > 0
                ORDER BY vote_count DESC
                LIMIT %s
            """, (limit,))
            
            top_voters = cursor.fetchall()
            return [dict(voter) for voter in top_voters]
        finally:
            cursor.close()
            conn.close()
    
    def create_announcement(self, title, content, language, created_by):
        """Yangi yangilik yaratish"""
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO announcements (title, content, language, created_by, created_at)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id
            """, (title, content, language, created_by, datetime.now()))
            
            announcement_id = cursor.fetchone()[0]
            conn.commit()
            return announcement_id
        except Exception as e:
            conn.rollback()
            print(f"Yangilik yaratishda xato: {e}")
            return None
        finally:
            conn.close()
    
    def get_last_announcement(self, language='uz'):
        """Oxirgi yangilikni olish"""
        conn = self.get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        try:
            cursor.execute("""
                SELECT title, content, created_at, created_by
                FROM announcements
                WHERE language = %s AND is_active = TRUE
                ORDER BY created_at DESC
                LIMIT 1
            """, (language,))
            
            result = cursor.fetchone()
            return dict(result) if result else None
        finally:
            cursor.close()
            conn.close()
    
    def get_settings_for_admin(self):
        """Admin uchun sozlamalar ro'yxati"""
        conn = self.get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        try:
            cursor.execute("""
                SELECT key, value, description, updated_at 
                FROM settings 
                ORDER BY key
            """)
            settings = cursor.fetchall()
            return [dict(setting) for setting in settings]
        finally:
            cursor.close()
            conn.close()
    
    def get_users_with_pending_balance(self):
        """Pending balansi bo'lgan foydalanuvchilarni olish"""
        conn = self.get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        try:
            cursor.execute("""
                SELECT id, telegram_id, first_name, last_name, pending_balance
                FROM users 
                WHERE pending_balance > 0 AND is_active = TRUE
                ORDER BY pending_balance DESC
            """)
            users = cursor.fetchall()
            return [dict(user) for user in users]
        finally:
            cursor.close()
            conn.close()
    
    def get_comprehensive_report_data(self):
        """To'liq hisobot uchun barcha ma'lumotlarni olish"""
        conn = None
        try:
            print("To'liq hisobot ma'lumotlari yig'ilmoqda...")
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # 1. Foydalanuvchilar va ularning ovozlari
            print("Foydalanuvchilar ma'lumotlari olinmoqda...")
            try:
                cursor.execute("""
                    SELECT 
                        u.id,
                        u.telegram_id,
                        u.username,
                        u.first_name,
                        u.last_name,
                        u.phone,
                        u.region,
                        u.language,
                        u.referral_code,
                        u.balance,
                        u.pending_balance,
                        u.total_earned,
                        u.total_withdrawn,
                        u.created_at,
                        COUNT(v.id) as total_votes,
                        COUNT(DISTINCT v.season_id) as seasons_voted
                    FROM users u
                    LEFT JOIN votes v ON u.id = v.user_id
                    WHERE u.is_active = TRUE
                    GROUP BY u.id
                    ORDER BY u.created_at DESC
                """)
                users_data = cursor.fetchall()
                print(f"Foydalanuvchilar ma'lumotlari olindi: {len(users_data)} ta")
            except Exception as e:
                print(f"Foydalanuvchilar ma'lumotlarini olishda xato: {e}")
                users_data = []
            
            # 2. Ovoz berish tarixi
            print("Ovoz berish tarixi olinmoqda...")
            try:
                cursor.execute("""
                    SELECT 
                        v.id,
                        v.user_id,
                        u.telegram_id,
                        u.first_name,
                        u.last_name,
                        p.name as project_name,
                        s.name as season_name,
                        v.created_at as vote_date
                    FROM votes v
                    JOIN users u ON v.user_id = u.id
                    JOIN projects p ON v.project_id = p.id
                    JOIN seasons s ON v.season_id = s.id
                    ORDER BY v.created_at DESC
                """)
                votes_data = cursor.fetchall()
                print(f"Ovoz berish tarixi olindi: {len(votes_data)} ta")
            except Exception as e:
                print(f"Ovoz berish tarixini olishda xato: {e}")
                votes_data = []
            
            # 3. Pul chiqarish so'rovlari
            print("Pul chiqarish so'rovlari olinmoqda...")
            try:
                cursor.execute("""
                    SELECT 
                        wr.id,
                        wr.user_id,
                        u.telegram_id,
                        u.first_name,
                        u.last_name,
                        u.phone,
                        wr.amount,
                        wr.commission,
                        wr.net_amount,
                        wr.method,
                        wr.account_details,
                        wr.status,
                        wr.created_at,
                        wr.processed_at
                    FROM withdrawal_requests wr
                    JOIN users u ON wr.user_id = u.id
                    ORDER BY wr.created_at DESC
                """)
                withdrawals_data = cursor.fetchall()
                print(f"Pul chiqarish so'rovlari olindi: {len(withdrawals_data)} ta")
            except Exception as e:
                print(f"Pul chiqarish so'rovlarini olishda xato: {e}")
                withdrawals_data = []
            
            # 4. Balans tarixi
            print("Balans tarixi olinmoqda...")
            try:
                cursor.execute("""
                    SELECT 
                        bh.id,
                        bh.user_id,
                        u.telegram_id,
                        u.first_name,
                        u.last_name,
                        bh.amount,
                        bh.type,
                        bh.description,
                        bh.status,
                        bh.created_at
                    FROM balance_history bh
                    JOIN users u ON bh.user_id = u.id
                    ORDER BY bh.created_at DESC
                """)
                balance_history_data = cursor.fetchall()
                print(f"Balans tarixi olindi: {len(balance_history_data)} ta")
            except Exception as e:
                print(f"Balans tarixini olishda xato: {e}")
                balance_history_data = []
            
            # 5. Loyihalar
            print("Loyihalar ma'lumotlari olinmoqda...")
            try:
                cursor.execute("""
                    SELECT 
                        p.id,
                        p.name,
                        p.budget,
                        p.region,
                        p.status,
                        s.name as season_name,
                        COUNT(v.id) as total_votes,
                        p.created_at
                    FROM projects p
                    LEFT JOIN seasons s ON p.season_id = s.id
                    LEFT JOIN votes v ON p.id = v.project_id
                    GROUP BY p.id, p.name, p.budget, p.region, p.status, s.name, p.created_at
                    ORDER BY p.created_at DESC
                """)
                projects_data = cursor.fetchall()
                print(f"Loyihalar ma'lumotlari olindi: {len(projects_data)} ta")
            except Exception as e:
                print(f"Loyihalar ma'lumotlarini olishda xato: {e}")
                projects_data = []
            
            # 6. Tasdiqlangan loyihalar
            print("Tasdiqlangan loyihalar olinmoqda...")
            try:
                cursor.execute("""
                    SELECT 
                        ap.id,
                        ap.name,
                        ap.link,
                        ap.status,
                        u.first_name as approved_by_name,
                        ap.approved_at,
                        ap.created_at
                    FROM approved_projects ap
                    JOIN users u ON ap.approved_by = u.id
                    ORDER BY ap.created_at DESC
                """)
                approved_projects_data = cursor.fetchall()
                print(f"Tasdiqlangan loyihalar olindi: {len(approved_projects_data)} ta")
            except Exception as e:
                print(f"Tasdiqlangan loyihalarni olishda xato: {e}")
                approved_projects_data = []
            
            print("Barcha ma'lumotlar muvaffaqiyatli yig'ildi!")
            
            return {
                'users': users_data,
                'votes': votes_data,
                'withdrawals': withdrawals_data,
                'balance_history': balance_history_data,
                'projects': projects_data,
                'approved_projects': approved_projects_data
            }
            
        except Exception as e:
            print(f"To'liq hisobot ma'lumotlarini olishda xato: {e}")
            print(f"Xato turi: {type(e).__name__}")
            import traceback
            print(f"Xato izi: {traceback.format_exc()}")
            
            # Bo'sh ma'lumotlar bilan qaytish
            return {
                'users': [],
                'votes': [],
                'withdrawals': [],
                'balance_history': [],
                'projects': [],
                'approved_projects': []
            }
            
        finally:
            if conn:
                conn.close()

    def create_excel_report(self):
        """Excel hisobot yaratish"""
        try:
            print("Excel hisobot yaratish boshlandi...")
            
            # openpyxl paketini tekshirish
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                print("openpyxl paketi muvaffaqiyatli import qilindi")
            except ImportError as e:
                print(f"openpyxl paketini import qilishda xato: {e}")
                return None
            
            # Ma'lumotlarni olish
            print("Ma'lumotlar yig'ilmoqda...")
            data = self.get_comprehensive_report_data()
            print(f"Ma'lumotlar yig'ildi: {len(data)} ta jadval")
            
            # Ma'lumotlar mavjudligini tekshirish
            total_records = sum(len(value) for value in data.values())
            if total_records == 0:
                print("Ma'lumotlar bazasi bo'sh - bo'sh Excel fayl yaratilmoqda...")
                # Bo'sh Excel fayl yaratish
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Ma'lumot yo'q"
                
                ws.cell(row=1, column=1, value="Ma'lumot yo'q")
                ws.cell(row=2, column=1, value="Ma'lumotlar bazasi bo'sh yoki ma'lumotlar topilmadi")
                ws.cell(row=3, column=1, value=f"Yaratilgan: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                
                filename = f"botopne_empty_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                wb.save(filename)
                print(f"Bo'sh Excel hisobot yaratildi: {filename}")
                return filename
            
            print(f"Jami {total_records} ta yozuv topildi, Excel fayl yaratilmoqda...")
            
            # Yangi Excel fayl yaratish
            print("Excel fayl yaratilmoqda...")
            wb = openpyxl.Workbook()
            
            # 1. Foydalanuvchilar sahifasi
            ws_users = wb.active
            ws_users.title = "Foydalanuvchilar"
            print("Foydalanuvchilar sahifasi yaratildi")
            
            # Sarlavhalar
            headers = [
                'ID', 'Telegram ID', 'Username', 'Ism', 'Familiya', 'Telefon', 
                'Hudud', 'Til', 'Referal kod', 'Balans', 'Pending', 
                'Jami topilgan', 'Jami chiqarilgan', 'Ro\'yxatdan o\'tish', 
                'Ovozlar soni', 'Mavsumlar'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws_users.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Ma'lumotlarni qo'shish
            print(f"Foydalanuvchilar ma'lumotlari qo'shilmoqda: {len(data['users'])} ta")
            for row, user in enumerate(data['users'], 2):
                ws_users.cell(row=row, column=1, value=user[0])  # ID
                ws_users.cell(row=row, column=2, value=user[1])  # Telegram ID
                ws_users.cell(row=row, column=3, value=user[2] or '')  # Username
                ws_users.cell(row=row, column=4, value=user[3] or '')  # First name
                ws_users.cell(row=row, column=5, value=user[4] or '')  # Last name
                ws_users.cell(row=row, column=6, value=user[5] or '')  # Phone
                ws_users.cell(row=row, column=7, value=user[6] or '')  # Region
                ws_users.cell(row=row, column=8, value=user[7] or '')  # Language
                ws_users.cell(row=row, column=9, value=user[8] or '')  # Referral code
                ws_users.cell(row=row, column=10, value=user[9] or 0)  # Balance
                ws_users.cell(row=row, column=11, value=user[10] or 0)  # Pending
                ws_users.cell(row=row, column=12, value=user[11] or 0)  # Total earned
                ws_users.cell(row=row, column=13, value=user[12] or 0)  # Total withdrawn
                ws_users.cell(row=row, column=14, value=str(user[13])[:19] if user[13] else '')  # Created at
                ws_users.cell(row=row, column=15, value=user[14] or 0)  # Total votes
                ws_users.cell(row=row, column=16, value=user[15] or 0)  # Seasons voted
            
            print("Foydalanuvchilar ma'lumotlari qo'shildi")
            
            # 2. Ovozlar sahifasi
            ws_votes = wb.create_sheet("Ovozlar")
            print("Ovozlar sahifasi yaratildi")
            
            vote_headers = ['ID', 'Foydalanuvchi ID', 'Telegram ID', 'Ism', 'Familiya', 'Loyiha', 'Mavsum', 'Sana']
            for col, header in enumerate(vote_headers, 1):
                cell = ws_votes.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row, vote in enumerate(data['votes'], 2):
                ws_votes.cell(row=row, column=1, value=vote[0])  # ID
                ws_votes.cell(row=row, column=2, value=vote[1])  # User ID
                ws_votes.cell(row=row, column=3, value=vote[2])  # Telegram ID
                ws_votes.cell(row=row, column=4, value=vote[3] or '')  # First name
                ws_votes.cell(row=row, column=5, value=vote[4] or '')  # Last name
                ws_votes.cell(row=row, column=6, value=vote[5] or '')  # Project name
                ws_votes.cell(row=row, column=7, value=vote[6] or '')  # Season name
                ws_votes.cell(row=row, column=8, value=str(vote[7])[:19] if vote[7] else '')  # Vote date
            
            print("Ovozlar ma'lumotlari qo'shildi")
            
            # 3. Pul chiqarish sahifasi
            ws_withdrawals = wb.create_sheet("Pul chiqarish")
            print("Pul chiqarish sahifasi yaratildi")
            
            withdrawal_headers = ['ID', 'Foydalanuvchi ID', 'Telegram ID', 'Ism', 'Familiya', 'Telefon', 'Miqdor', 'Komissiya', 'Olinadigan', 'Usul', 'Ma\'lumotlar', 'Holat', 'Sana', 'Tasdiqlangan']
            for col, header in enumerate(withdrawal_headers, 1):
                cell = ws_withdrawals.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row, withdrawal in enumerate(data['withdrawals'], 2):
                ws_withdrawals.cell(row=row, column=1, value=withdrawal[0])  # ID
                ws_withdrawals.cell(row=row, column=2, value=withdrawal[1])  # User ID
                ws_withdrawals.cell(row=row, column=3, value=withdrawal[2])  # Telegram ID
                ws_withdrawals.cell(row=row, column=4, value=withdrawal[3] or '')  # First name
                ws_withdrawals.cell(row=row, column=5, value=withdrawal[4] or '')  # Last name
                ws_withdrawals.cell(row=row, column=6, value=withdrawal[5] or '')  # Phone
                ws_withdrawals.cell(row=row, column=7, value=withdrawal[6] or 0)  # Amount
                ws_withdrawals.cell(row=row, column=8, value=withdrawal[7] or 0)  # Commission
                ws_withdrawals.cell(row=row, column=9, value=withdrawal[8] or 0)  # Net amount
                ws_withdrawals.cell(row=row, column=10, value=withdrawal[9] or '')  # Method
                ws_withdrawals.cell(row=row, column=11, value=withdrawal[10] or '')  # Account details
                ws_withdrawals.cell(row=row, column=12, value=withdrawal[11] or '')  # Status
                ws_withdrawals.cell(row=row, column=13, value=str(withdrawal[12])[:19] if withdrawal[12] else '')  # Created at
                ws_withdrawals.cell(row=row, column=14, value=str(withdrawal[13])[:19] if withdrawal[13] else '')  # Processed at
            
            print("Pul chiqarish ma'lumotlari qo'shildi")
            
            # 4. Balans tarixi sahifasi
            ws_balance = wb.create_sheet("Balans tarixi")
            print("Balans tarixi sahifasi yaratildi")
            
            balance_headers = ['ID', 'Foydalanuvchi ID', 'Telegram ID', 'Ism', 'Familiya', 'Miqdor', 'Turi', 'Izoh', 'Holat', 'Sana']
            for col, header in enumerate(balance_headers, 1):
                cell = ws_balance.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row, balance in enumerate(data['balance_history'], 2):
                ws_balance.cell(row=row, column=1, value=balance[0])  # ID
                ws_balance.cell(row=row, column=2, value=balance[1])  # User ID
                ws_balance.cell(row=row, column=3, value=balance[2])  # Telegram ID
                ws_balance.cell(row=row, column=4, value=balance[3] or '')  # First name
                ws_balance.cell(row=row, column=5, value=balance[4] or '')  # Last name
                ws_balance.cell(row=row, column=6, value=balance[5] or 0)  # Amount
                ws_balance.cell(row=row, column=7, value=balance[6] or '')  # Type
                ws_balance.cell(row=row, column=8, value=balance[7] or '')  # Description
                ws_balance.cell(row=row, column=9, value=balance[8] or '')  # Status
                ws_balance.cell(row=row, column=10, value=str(balance[9])[:19] if balance[9] else '')  # Created at
            
            print("Balans tarixi ma'lumotlari qo'shildi")
            
            # 5. Loyihalar sahifasi
            ws_projects = wb.create_sheet("Loyihalar")
            print("Loyihalar sahifasi yaratildi")
            
            project_headers = ['ID', 'Nomi', 'Budjet', 'Hudud', 'Holat', 'Mavsum', 'Ovozlar soni', 'Yaratilgan']
            for col, header in enumerate(project_headers, 1):
                cell = ws_projects.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row, project in enumerate(data['projects'], 2):
                ws_projects.cell(row=row, column=1, value=project[0])  # ID
                ws_projects.cell(row=row, column=2, value=project[1] or '')  # Name
                ws_projects.cell(row=row, column=3, value=project[2] or 0)  # Budget
                ws_projects.cell(row=row, column=4, value=project[3] or '')  # Region
                ws_projects.cell(row=row, column=5, value=project[4] or '')  # Status
                ws_projects.cell(row=row, column=6, value=project[5] or '')  # Season name
                ws_projects.cell(row=row, column=7, value=project[6] or 0)  # Total votes
                ws_projects.cell(row=row, column=8, value=str(project[7])[:19] if project[7] else '')  # Created at
            
            print("Loyihalar ma'lumotlari qo'shildi")
            
            # 6. Tasdiqlangan loyihalar sahifasi
            ws_approved = wb.create_sheet("Tasdiqlangan loyihalar")
            print("Tasdiqlangan loyihalar sahifasi yaratildi")
            
            approved_headers = ['ID', 'Nomi', 'Havola', 'Holat', 'Tasdiqlagan', 'Tasdiqlangan', 'Yaratilgan']
            for col, header in enumerate(approved_headers, 1):
                cell = ws_approved.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row, approved in enumerate(data['approved_projects'], 2):
                ws_approved.cell(row=row, column=1, value=approved[0])  # ID
                ws_approved.cell(row=row, column=2, value=approved[1] or '')  # Name
                ws_approved.cell(row=row, column=3, value=approved[2] or '')  # Link
                ws_approved.cell(row=row, column=4, value=approved[3] or '')  # Status
                ws_approved.cell(row=row, column=5, value=approved[4] or '')  # Approved by
                ws_approved.cell(row=row, column=6, value=str(approved[5])[:19] if approved[5] else '')  # Approved at
                ws_approved.cell(row=row, column=7, value=str(approved[6])[:19] if approved[6] else '')  # Created at
            
            print("Tasdiqlangan loyihalar ma'lumotlari qo'shildi")
            
            # Sütunlarni avtomatik o'lchamlash
            print("Sütunlarni avtomatik o'lchamlash...")
            for ws in [ws_users, ws_votes, ws_withdrawals, ws_balance, ws_projects, ws_approved]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            print("Sütunlar o'lchamlandi")
            
            # Excel faylni saqlash
            filename = f"botopne_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            print(f"Excel fayl saqlanmoqda: {filename}")
            wb.save(filename)
            print(f"Excel hisobot muvaffaqiyatli yaratildi: {filename}")
            
            return filename
            
        except ImportError as e:
            print(f"openpyxl paketi topilmadi: {e}")
            print("Excel hisobot yaratish uchun: pip install openpyxl")
            return None
        except Exception as e:
            print(f"Excel hisobot yaratishda xato: {e}")
            print(f"Xato turi: {type(e).__name__}")
            import traceback
            print(f"Xato izi: {traceback.format_exc()}")
            return None
